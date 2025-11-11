"""
skills_extractor.py
====================

This module provides a command‑line utility to extract skills from an Excel
workbook, consolidate them into a unique, sorted list, and generate a
review worksheet for the client. It can also pre‑mark skills that
originate from the first column (typically LinkedIn skills) and format
the output as an Excel table with alternating row colours for easier
reading.

Usage
-----

Run the script from the command line, supplying at least the path to
the source workbook. Optionally specify a sheet name to read from,
the output path for the generated workbook, and the client’s name
for use in the default share message.

Examples:

    python skills_extractor.py client_skills.xlsx

    python skills_extractor.py client_skills.xlsx --sheet "October" --output
        client_skills_review.xlsx --client "Morgan"

The generated workbook will contain a sheet called ``Skills Review`` with
two columns: ``Skill`` and ``Have skill? (Mark X or YES)``. Skills found
in the first column of the input sheet are automatically marked with
``Yes`` in the second column.

Dependencies
------------

This script relies on `pandas` for data handling and `openpyxl` for
creating Excel files. Make sure both packages are installed in your
environment.

"""

import argparse
from pathlib import Path
from typing import List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def flatten_values(values: List[str]) -> List[str]:
    """Split values by commas and newlines and strip whitespace."""
    flattened: List[str] = []
    for val in values:
        if not isinstance(val, str):
            val = str(val)
        for part in val.replace("\n", ",").split(","):
            part = part.strip()
            if part:
                flattened.append(part)
    return flattened


def extract_unique_skills(
    excel_path: Path, sheet: Optional[str] = None
) -> List[str]:
    """Read all cells from the specified sheet and return a sorted list of unique skills."""
    xls = pd.ExcelFile(excel_path)
    sheet_names = [sheet] if sheet else xls.sheet_names
    all_skills: List[str] = []
    for sheet_name in sheet_names:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        for col in df.columns:
            col_values = df[col].dropna().tolist()
            all_skills.extend(flatten_values(col_values))
    # Deduplicate and sort (case‑insensitive)
    unique_skills = sorted(set(all_skills), key=lambda s: s.lower())
    return unique_skills


def determine_linkedin_skills(
    df: pd.DataFrame
) -> set:
    """Return a set of skills listed in the first column of the DataFrame."""
    first_col = df.columns[0]
    values = df[first_col].dropna().tolist()
    return set(flatten_values(values))


def create_review_workbook(
    skills: List[str],
    linkedin_skills: set,
    output_path: Path,
    table_style_name: str = "TableStyleMedium9",
    show_row_stripes: bool = True,
    show_column_stripes: bool = False,
) -> None:
    """
    Create an Excel workbook with the consolidated skill list and
    prefilled LinkedIn skills, applying a table style for alternating colours.

    Parameters
    ----------
    skills : List[str]
        The list of unique skills to include.
    linkedin_skills : set
        A set of skills that should be pre‑marked as "Yes".
    output_path : Path
        Where to save the resulting workbook.
    table_style_name : str, optional
        The name of the Excel table style to use. Defaults to 'TableStyleMedium9'.
    show_row_stripes : bool, optional
        Whether to show alternating row colours.
    show_column_stripes : bool, optional
        Whether to show alternating column colours.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Skills Review"
    # Write header
    ws["A1"] = "Skill"
    ws["B1"] = "Have skill? (Mark X or YES)"
    # Write skills and prefill LinkedIn skills
    for idx, skill in enumerate(skills, start=2):
        ws[f"A{idx}"] = skill
        if skill in linkedin_skills:
            ws[f"B{idx}"] = "Yes"
    # Adjust column widths
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 30

    # Freeze the first row so headers remain visible when scrolling
    # Freeze panes at the first cell below the header row (A2)
    ws.freeze_panes = "A2"
    # Define table range
    end_row = len(skills) + 1
    table_ref = f"A1:B{end_row}"
    table = Table(displayName="SkillsTable", ref=table_ref)
    # Apply style with alternating colours
    style = TableStyleInfo(
        name=table_style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=show_row_stripes,
        showColumnStripes=show_column_stripes,
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    wb.save(output_path)


def build_share_message(client_name: str, sender_name: str) -> str:
    """Return a sample share message for the client."""
    return (
        f"Hi {client_name},\n\n"
        "I've compiled a list of possible job titles and their skills for you."
        " Your current LinkedIn skills are already marked."
        " Please review the list and mark 'Yes' in the second column for any"
        " additional skills you have.\n\n"
        "Thanks,\n"
        f"{sender_name}"
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extract unique skills from an Excel workbook and create a review sheet."
    )
    parser.add_argument(
        "excel_path",
        type=Path,
        help="Path to the source Excel file containing skill lists.",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Name of the worksheet to process. Defaults to the first sheet.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help=(
            "Path to the output Excel file. If not provided, a new file"
            " with suffix '_review' will be created next to the input file."
        ),
    )
    parser.add_argument(
        "--client",
        type=str,
        default="Client",
        help="Name of the client for the share message.",
    )
    parser.add_argument(
        "--sender",
        type=str,
        default="Your Name",
        help="Your name to include in the share message.",
    )
    args = parser.parse_args()

    excel_path = args.excel_path
    if not excel_path.exists():
        parser.error(f"Input file not found: {excel_path}")

    # Determine output path
    if args.output is None:
        output_path = excel_path.with_name(excel_path.stem + "_review.xlsx")
    else:
        output_path = args.output

    # Read input data
    df = pd.read_excel(excel_path, sheet_name=args.sheet)
    linkedin_skills = determine_linkedin_skills(df)
    unique_skills = extract_unique_skills(excel_path, sheet=args.sheet)

    # Generate review workbook
    create_review_workbook(unique_skills, linkedin_skills, output_path)

    # Print out share message for convenience
    message = build_share_message(args.client, args.sender)
    print("Review workbook created at:", output_path)
    print("\nSuggested share message:\n")
    print(message)


def process_skills(
    excel_path: str,
    sheet: Optional[str] = None,
    client: str = "Client",
    sender: str = "Your Name",
) -> dict:
    """
    Process an Excel workbook and generate a review sheet.

    Parameters
    ----------
    excel_path : str
        Path to the source Excel file.
    sheet : str, optional
        Name of the worksheet to process. Defaults to the first sheet.
    client : str, optional
        The client’s name for the share message.
    sender : str, optional
        Your name for the share message.

    Returns
    -------
    dict
        A dictionary with keys ``output_file`` and ``message`` containing the path
        to the generated review workbook and a suggested share message.
    """
    src = Path(excel_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")
    # Create output path with suffix _review.xlsx in the same directory
    output_path = src.with_name(src.stem + "_review.xlsx")
    df = pd.read_excel(src, sheet_name=sheet)
    linkedin_skills = determine_linkedin_skills(df)
    unique_skills = extract_unique_skills(src, sheet=sheet)
    create_review_workbook(unique_skills, linkedin_skills, output_path)
    message = build_share_message(client, sender)
    return {"output_file": str(output_path), "message": message}


if __name__ == "__main__":
    main()